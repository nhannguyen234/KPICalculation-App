"""
Created by Nguyen Le Nhan 
Updated date: 29 March 2023
"""
import os 
import re
import numpy as np
import openpyxl as xl
import pandas as pd
from tkinter import messagebox

def find_sensor(df_path):
    combined_string = ''.join(df_path)
    pattern = r'M[-_]*\d+[-_]*\d+[-_]*\d*'
    empty, underscore = '', '_'
    matches = re.findall(pattern, combined_string)
    # replace characters to the consistent format
    subs_string_list = list(dict.fromkeys(['M' + re.sub(r'M[-_]*', empty, s) for s in matches]))
    subs_string_list = [re.sub(r'[\-]+', underscore, s) for s in subs_string_list]
    return subs_string_list

# def seq_column(df):
#     """Make columns in Alphabet sequence"""
#     list_df = list(df.columns[1:])
#     list_df.sort()
#     list_df.insert(0,'f[Hz]')
#     return list_df

def extract_TF_frommeas(merge,
                        files, 
                        name, 
                        cols_name,
                        complex_component_list):
    """
        Function: Extract transfer function from measurement database
        Format Excel file: based on the format of Excel from Prashan
    """
    global magnitude, real, imaginary
    
    col_name_list = ['XX', 'XY', 'XZ','YX','YY','YZ','ZX','ZY','ZZ']

    files_path = [files.split(' ')[i] for i in range(len(files.split(' ')))]
    files_path.sort()
    cols_list = [[i, i+1] for i in range(2,20,2)]
    
    dict_cols = {col_name_list[i]:cols_list[i] for i in range(len(col_name_list))}
    series = pd.Series(np.linspace(100,2000,1901))
 
    for col_name in cols_name:
        csv_file_list =[]
        direction = col_name
        cols = dict_cols[col_name]
        for file in files_path:
            wb1 = xl.load_workbook(file)
            ws1 = wb1.worksheets[0]
            # calculate total number of rows and 
            # columns in source excel file
            mr = ws1.max_row
            magnitude = []
            real = []
            imaginary = []
            for r in range(3,mr+1):
                real_val = float(ws1.cell(row = r, column=cols[0]).value)
                img_val = float(ws1.cell(row = r, column=cols[1]).value)
                real.append(real_val)
                imaginary.append(img_val)
                magnitude.append(np.sqrt(pow(real_val,2) + pow(img_val,2)))

            for complex_component in complex_component_list:
                sensor_name =  "_".join(find_sensor(file))
                globals()['df_' + complex_component + sensor_name] = pd.DataFrame(list(zip(series, globals()[str(complex_component)])), columns=['f[Hz]',sensor_name])

        for complex_component in complex_component_list:
            df_final = globals()['df_' + complex_component + "_".join(find_sensor(files_path[0]))]
            for i in range(1,len(files_path)):
                df_final = df_final.merge(globals()['df_'  + complex_component + "_".join(find_sensor(files_path[i]))], how='inner', on='f[Hz]')
            
            try:
                new_file_path = os.path.join(os.path.dirname(file),
                                            name +'_' + direction + '_measurement_converted' + '_' + complex_component + '.csv')
                csv_file_list.append(new_file_path)
                df_final.to_csv((new_file_path), index=False)
            except:
                messagebox.showerror(title='Error', 
                                     message='Please close file ' + os.path.join(os.path.dirname(file), 
                                     name +'_' + direction + '_measurement_converted' + '_' + complex_component + '.csv'))
                raise Exception
        if merge == 1:
            merged_csv(csv_file_list)


def expand_data_SIM(files):
    """
        Function: Convert Abaqus's Excel format to csv files
        Format Excel file: based on the format of Excel exported from Abaqus plugins
    """
    files_path = [files.split(' ')[i] for i in range(len(files.split(' ')))]
    def expand_data(init_num, last_num, init_val, last_val): #linear regression data
        inserted_data=[]
        delta = last_num - init_num
        for i in range(delta-1):
            inserted_data.append(init_val + (i+1)*(last_val-init_val)/delta)
        return inserted_data

    for file in files_path:
        wb1 = xl.load_workbook(file)
        ws1 = wb1.worksheets[0]
        # calculate total number of rows and 
        # columns in source excel file
        mr = ws1.max_row
        mc = ws1.max_column
        for c in range(4,mc+1):
            try:
                new_transferfunction=[]
                for r in range(14,mr):
                    new_transferfunction.append(float(ws1.cell(row = r, column=c).value))
                    list_expand = expand_data(int(ws1.cell(row = r, column=3).value),\
                                              int(ws1.cell(row = r+1, column=3).value),\
                                              float(ws1.cell(row = r, column=c).value),\
                                              float(ws1.cell(row = r+1, column=c).value))
                    for i in range(len(list_expand)):
                        new_transferfunction.append(list_expand[i])
                new_transferfunction.append(float(ws1.cell(row = mr, column=c).value))
                series = pd.Series(np.linspace(100,2000,1901))
                globals()['df_' + str(ws1.cell(row = 6, column=c).value)] = pd.DataFrame(list(zip(series, new_transferfunction)), columns=['f[Hz]',str(ws1.cell(row = 6, column=c).value)])
            except:
                continue
        df_final = globals()['df_' + str(ws1.cell(row = 6, column=4).value)]
        for i in range(5,mc+1):
            try:
                df_final = df_final.merge(globals()['df_' + str(ws1.cell(row = 6, column=i).value)], how='inner', on='f[Hz]')
            except:
                continue
        try:
            df_final.to_csv(file.split('.')[0] + '_sim_converted.csv', index=False)
        except:
            messagebox.showerror(title='Error', message='Please close file ' + file.split('.')[0] + '_sim_converted.csv')
            raise Exception

def merged_csv(csv_file_list):
    """Merge all components data to one file name with suffixed 'merged' """
    files_path = csv_file_list
    specific_order = ['real', 'imaginary', 'magnitude']
    order = {key: i for i, key in enumerate(specific_order)}
    #create a new list with specific order
    new_file_paths = sorted(files_path, key=lambda files_path: order[files_path.split('_')[-1].split('.')[0]])

    #inititate a dataframe for merge dataframe
    df_final = pd.read_csv(new_file_paths[0])\
                 .set_index(['f[Hz]'])\
                 .add_suffix('_' + os.path.basename(new_file_paths[0]).split('_')[-1].split('.')[0])

    for i in range(1,len(new_file_paths)):
        globals()['df_' + os.path.basename(new_file_paths[i]).split('.')[0]] = pd.read_csv(new_file_paths[i])\
                                                                                  .set_index(['f[Hz]'])\
                                                                                  .add_suffix('_' + os.path.basename(new_file_paths[i]).split('_')[-1].split('.')[0])
        df_final = df_final.merge(globals()['df_' + os.path.basename(new_file_paths[i]).split('.')[0]], 
                                  how='inner',
                                  on='f[Hz]')
    try:
        df_final.to_csv('_'.join(new_file_paths[0].split('.')[0].split('_')[:-1]) + '_merged.csv', index=False)
    except:
        messagebox.showerror(title='Error', 
                            message='Please close file ' + '_'.join(new_file_paths[0].split('.')[0].split('_')[:-1]) + '_merged.csv')
        raise Exception
    
